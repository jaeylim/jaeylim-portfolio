"""
Step 1: 엑셀 체크리스트 -> 구조화된 JSON

원본 엑셀은 병합 셀(범주/분야/통제항목)을 쓰고 있어서,
그대로 읽으면 대부분 None이 나옴. merged_cells 정보를 이용해
각 행에 값을 forward-fill 해준다.
"""
import json
import openpyxl

SRC = "checklists.xlsx"  # 실제 파일명으로 바꿔서 실행
SHEET = "checksheet"  # 실제 시트명으로 바꿔서 실행
OUT = "checklist.json"


def load_with_merged_fill(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    # 병합 범위: {(row, col): (value, min_row, max_row, min_col, max_col)}
    merge_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(r, c)] = top_left

    def cell_value(r, c):
        if (r, c) in merge_map:
            return merge_map[(r, c)]
        return ws.cell(r, c).value

    rows = []
    # 데이터는 3행부터 시작 (1행 공백, 2행 헤더)
    for r in range(3, ws.max_row + 1):
        category = cell_value(r, 2)      # B: 범주
        domain = cell_value(r, 3)        # C: 분야
        control_item = cell_value(r, 4)  # D: 통제항목 (코드 포함)
        evidence = cell_value(r, 5)      # E: 필요 증적자료
        note = cell_value(r, 6)          # F: 비고

        if not evidence:
            continue

        rows.append({
            "row": r,
            "category": category,
            "domain": domain,
            "control_item": control_item,
            "evidence_required": evidence,
            "note": note or "",
        })
    return rows


if __name__ == "__main__":
    rows = load_with_merged_fill(SRC, SHEET)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"{len(rows)}개 항목을 {OUT} 에 저장했습니다.")
